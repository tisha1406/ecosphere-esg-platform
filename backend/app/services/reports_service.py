import io
from typing import Dict, Any
from uuid import UUID
from sqlalchemy.ext.asyncio import AsyncSession
from app.services import scoring_service
from app.schemas.reports import ConsolidatedReport, TrendPoint, InitiativeSummary, ContributorSummary
from sqlalchemy import select, desc
from app.models.social import CsrInitiative
from app.models.user import User
from app.models.gamification import PointsHistory
from sqlalchemy.sql import func
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class ReportsService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_consolidated_report(self, company_id: str, period: str) -> ConsolidatedReport:
        # 1. Get Scores
        summary = await scoring_service.get_score_summary(self.db, company_id, period)
        
        # 2. Get Top Initiatives (from Social)
        result_init = await self.db.execute(
            select(CsrInitiative).where(CsrInitiative.is_active == True).limit(5)
        )
        initiatives = result_init.scalars().all()
        top_initiatives = [
            InitiativeSummary(name=i.title, impact=f"Budget: ${i.budget}")
            for i in initiatives
        ]

        # 3. Get Top Contributors (from Gamification)
        # Assuming gamification points ledger or PointsHistory
        # We will aggregate points per user
        result_points = await self.db.execute(
            select(User.full_name, func.sum(PointsHistory.points).label("total_points"))
            .join(PointsHistory, User.id == PointsHistory.user_id)
            .group_by(User.id)
            .order_by(desc("total_points"))
            .limit(5)
        )
        top_contributors = [
            ContributorSummary(name=row[0] or "Unknown User", points=row[1] or 0)
            for row in result_points.all()
        ]
        
        # 4. Trend Series (Dummy for now as time series score tracking might not exist)
        trend_series = [
            TrendPoint(month="Jan", score=70),
            TrendPoint(month="Feb", score=72),
            TrendPoint(month="Mar", score=75),
            TrendPoint(month="Apr", score=summary.total_score or 75),
        ]

        return ConsolidatedReport(
            company_name="EcoSphere Client", # In reality, fetch from Company model
            period=period,
            environmental_score=summary.environmental_score,
            social_score=summary.social_score,
            governance_score=summary.governance_score,
            overall_esg_score=summary.total_score,
            trend_series=trend_series,
            top_initiatives=top_initiatives,
            top_contributors=top_contributors
        )

    async def generate_pdf_report(self, report_data: ConsolidatedReport) -> io.BytesIO:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']
        
        elements.append(Paragraph(f"ESG Consolidated Report: {report_data.company_name}", title_style))
        elements.append(Paragraph(f"Period: {report_data.period}", normal_style))
        elements.append(Spacer(1, 12))
        
        # Scores
        elements.append(Paragraph("Overall Scores", heading_style))
        data = [
            ["Category", "Score"],
            ["Environmental", f"{report_data.environmental_score}"],
            ["Social", f"{report_data.social_score}"],
            ["Governance", f"{report_data.governance_score}"],
            ["Overall ESG", f"{report_data.overall_esg_score}"]
        ]
        t = Table(data, colWidths=[200, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor("#2e7d32")),
            ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 24))
        
        # Initiatives
        elements.append(Paragraph("Top CSR Initiatives", heading_style))
        init_data = [["Initiative Name", "Impact"]]
        for init in report_data.top_initiatives:
            init_data.append([init.name, init.impact])
            
        t2 = Table(init_data, colWidths=[200, 200])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor("#1565c0")),
            ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        elements.append(t2)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    async def generate_xlsx_report(self, report_data: ConsolidatedReport) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESG Report"
        
        # Headers
        ws["A1"] = "EcoSphere ESG Report"
        ws["A2"] = f"Company: {report_data.company_name}"
        ws["A3"] = f"Period: {report_data.period}"
        
        ws["A5"] = "Scores"
        ws["A6"] = "Environmental"
        ws["B6"] = report_data.environmental_score
        ws["A7"] = "Social"
        ws["B7"] = report_data.social_score
        ws["A8"] = "Governance"
        ws["B8"] = report_data.governance_score
        ws["A9"] = "Overall ESG"
        ws["B9"] = report_data.overall_esg_score
        
        ws["A11"] = "Top CSR Initiatives"
        ws["A12"] = "Name"
        ws["B12"] = "Impact"
        row = 13
        for init in report_data.top_initiatives:
            ws[f"A{row}"] = init.name
            ws[f"B{row}"] = init.impact
            row += 1
            
        ws[f"A{row+1}"] = "Top Contributors"
        ws[f"A{row+2}"] = "Name"
        ws[f"B{row+2}"] = "Points"
        row += 3
        for cont in report_data.top_contributors:
            ws[f"A{row}"] = cont.name
            ws[f"B{row}"] = cont.points
            row += 1
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
